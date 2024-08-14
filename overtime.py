import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment

adi_sort_key = {
    "署長室": 1, "胡副署長室": 2, "林副署長室": 3, "主任秘書室": 4,
    "政策規劃組": 5, "前瞻政策科": 6, "產業人才科": 7, "計畫管理科": 8, "綜合業務科": 9,
    "通訊傳播組": 10, "基礎環境科": 11, "傳播推廣科": 12, "通訊應用科": 13,
    "平臺經濟組": 14, "平臺應用科": 15, "平臺治理科": 16, "數位體驗科": 17, "數據經濟科": 18,
    "新興跨域組": 19, "資安產業科": 20, "全球運籌科": 21, "場域實證科": 22, "地方鏈結科": 23,
    "數位服務組": 24, "軟體產業科": 25, "轉型輔導科": 26, "整合服務科": 27, "創新應用科": 28,
    "秘書室": 29, "文書科": 30, "事務科": 31,
    "人事室": 32, "政風室": 33, "主計室": 34,
    "署長": 35, "副署長": 36, "主任秘書": 37,
    "組長": 38, "副組長": 39, "簡任技正": 40, "簡任視察": 41,
    "科長": 42, "技正": 43, "視察": 44, "專員": 45, "科員": 46, "助理員": 47,
    "專案規劃師": 48, "專案分析師": 49
}


def clean_data(source_df: pd.DataFrame):
    """
    :param source_df: source dataframe
    :return: reformatted dataframe
    """
    ch_cols = ["單位名稱", "職稱", "姓名", "加班日期", "核可時數", "已請款時數", "已補休時數", "剩餘可用時數"]
    en_cols = ["unit", "job", "name", "mons", "appr", "paid", "rest", "remain"]
    result: pd.DataFrame = source_df.copy(deep=True)
    result.columns = source_df.iloc[3]
    result = result.iloc[4:][ch_cols].copy(deep=True)
    result.rename(columns={ch_cols[i]: en_cols[i] for i in range(len(ch_cols))}, inplace=True)
    result["mons"] = result["mons"].str.split("/").map(lambda x: str(int(x[1])))
    # split hours and minutes
    for col in en_cols[4:]:
        result[col] = result.fillna("0-0")
        result[f"{col}_H"] = result[col].str.split("-", expand=True).loc[:, 0].astype(int)
        result[f"{col}_M"] = result[col].str.split("-", expand=True).loc[:, 1].astype(int)

    result = result.drop(en_cols[4:], axis="columns")

    return result


def extract_mons_data(df: pd.DataFrame):
    result, mons = {}, df["mons"].unique()
    for mon in mons:
        inner_df = df[df["mons"] == mon]
        for col in inner_df.columns[4:]:
            result = {**result, **{f"{col}_{mon}": inner_df[col].sum()}}

    return result


def sorted_column_dict(col_patterns) -> dict:
    columns, result = [], {}
    for i in range(1, 13):
        for pattern in col_patterns:
            columns.append(f"{pattern}_H_{str(i)}")
            columns.append(f"{pattern}_M_{str(i)}")

    for i, col in enumerate(columns):
        result[col] = i

    return result


def integrate_hours_mins(df):
    result = df.fillna(0).copy(deep=True)
    # the columns for hours and minutes which need to be integrated.
    col_patterns = ["appr", "paid", "rest", "remain"]
    mons = set([col.split("_")[2] for col in result.columns[3:]])
    for mon in mons:
        for pattern in col_patterns:
            hour_col, min_col = f"{pattern}_H_{mon}", f"{pattern}_M_{mon}"
            result[hour_col] = result[hour_col] + (result[min_col] / 60).astype(int)
            result[min_col] = (result[min_col] % 60).astype(int)

    sort_dict = sorted_column_dict(col_patterns)
    result = result[list(result.columns[:3]) + sorted(result.columns[3:], key=lambda v: sort_dict.get(v))]

    return result


def aggregate_time(source_df):
    aggregate_data = {}
    for name in source_df["name"]:
        df = source_df[source_df["name"] == name]
        aggregate_data[name] = {
            "unit": df["unit"].values[0], "job": df["job"].values[0], "name": df["name"].values[0]}
        aggregate_data[name] = {**aggregate_data[name], **extract_mons_data(df)}

    result = pd.DataFrame.from_dict(data=aggregate_data, orient="index")
    result.reset_index().rename(columns={"index": "name"}, inplace=True)
    # the columns for hours and minutes which need to be integrated.
    result = integrate_hours_mins(result)
    for col in result.columns[3:]:
        result[col] = result[col].astype(int)

    return result


# excel part
def applying_style(ws: Workbook.worksheets, num_rows, num_cols, attributes):
    if ws.title in attributes:
        ws["A1"] = "idx"
        ws.delete_rows(2)
    for row in range(1, num_rows + 2):
        for col in range(1, num_cols + 2):
            ws.cell(row=row, column=col).font = Font(name="Microsoft JhengHei", size=12)
            ws.cell(row=row, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')


def create_general_sheet(source_df, attributes):
    wb = Workbook()
    for attribute in attributes:
        if attribute == "resource":
            ws = wb.active
            ws.title = "resource"
            df = source_df.copy(deep=True)
        else:
            cols = [col for col in source_df.columns if attribute in col]
            ws = wb.create_sheet(attribute)
            df = source_df[list(source_df.columns[:3]) + cols].copy(deep=True)

        for r in dataframe_to_rows(df, index=True):
            ws.append(r)

        applying_style(ws, num_rows=df.shape[0], num_cols=df.shape[1], attributes=attributes)

    return wb


def create_single_sheet(source_df: pd.DataFrame, wb: Workbook, names: list, attributes: list):
    # noinspection PyShadowingNames
    def insert_metadata(ws, df):
        for col in reversed(source_df.columns[:3]):
            ws.insert_cols(1)
            ws["A1"] = df.loc[0, col]

        ws.merge_cells("A1:A8")
        ws.merge_cells("B1:B8")
        ws.merge_cells("C1:C8")

    for name in names:
        # print(source_df["name"])
        if name in list(source_df["name"]):
            ws: Workbook.worksheets = wb.create_sheet(name)
            df = source_df[source_df["name"] == name].copy(deep=True)
            df.reset_index(drop=True, inplace=True)
            i_row, cols = 1, []
            for attribute in attributes[1:]:
                cols = [col for col in source_df.columns if attribute in col]
                for j_col, col in enumerate(cols):
                    ws.cell(row=i_row, column=j_col + 1).value = col
                    ws.cell(row=i_row + 1, column=j_col + 1).value = df.loc[0, col]

                i_row += 2

            insert_metadata(ws, df)
            applying_style(ws, num_rows=7, num_cols=len(cols) + 2, attributes=attributes)
        else:
            print(f"{name} 不在你家上班~")


def to_xlxs_file(source_df: pd.DataFrame, names: list):
    attributes = ["resource", "appr", "paid", "rest", "remain"]
    wb = create_general_sheet(source_df, attributes)
    if names:
        create_single_sheet(source_df, wb, names, attributes)

    wb.save("超時服勤時數統計表.xlsx")


# main execution
def overtime_statistics(file_name, sort_key, names=None):
    if names is None:
        names = []
    df = pd.read_excel(file_name)
    result = aggregate_time(clean_data(df))
    result = result.sort_values(by=["unit", "job"], key=lambda x: x.map(sort_key))
    result = result.reset_index(drop=True)
    to_xlxs_file(result, names)

    return result
